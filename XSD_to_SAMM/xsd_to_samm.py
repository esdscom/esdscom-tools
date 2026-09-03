#!/usr/bin/env python3
"""
xsd_to_samm.py
==============
Converts the eSDScom/sdscom-xml XSD schema to a SAMM 2.1.0 Turtle ontology
for validation in the Catena-X Aspect Model Editor.

Issues found in SAMM-Beispiel.ttl – corrected in this script:
  [ERROR]   samm:preferredName replaces the non-existent samm:name property;
            in SAMM 2.1.0 the element name is the IRI local fragment only
  [ERROR]   samm:minCount → not a SAMM property; optional properties belong
            in samm:properties as [ samm:property :p ; samm:optional true ]
  [ERROR]   ". ;" → invalid Turtle; trailing dot already terminates a statement
  [WARNING] samm:Characteristic for entity refs → use samm-c:SingleEntity
  [HINT]    @prefix samm-u: → standard SAMM prefix is "unit:"
  [HINT]    .../link/SDScomChem → user spec: https://esdscom.eu/SDScomChem

CamelCase conventions (SAMM / Catena-X):
  Property IRIs     : lowerCamelCase     (dataSheet, specificationNo)
  Entity/Aspect IRIs: PascalCase         (DataSheet, DatasheetFeed)
  Acronym fix       : SDSComXML→SdsComXml, ClassificationEU→ClassificationEu
  Deduplication     : when same element name has multiple types, the type's
                      local name is appended as suffix to the property IRI
                      (classificationClassificationEu vs …De)

Usage
-----
    python xsd_to_samm.py [xsd_dir] [output_ttl]
    xsd_dir     – directory for XSD files (default ./xsd_files, auto-downloaded)
    output_ttl  – output filename          (default DatasheetFeed.ttl)
"""

import os, sys, re, urllib.request
from typing import Optional, Set
from xml.etree import ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

GITHUB_RAW    = "https://raw.githubusercontent.com/esdscom/sdscom-xml/master/"
SAMM_VERSION  = "2.1.0"
MODEL_NS      = "urn:samm:io.catenax.esdscom:1.0.0#"
APPINFO_BASE  = "https://esdscom.eu/"      # samm:see base URL for appinfo-derived links
ASPECT_IRI    = "Esdscom"
ASPECT_SEE_URL = "https://esdscom.eu/esdscom_5_6"

# Local type names whose child elements are NOT expanded (stub entities created)
STOP_TYPES: set = set()

# ── REUSED from xsd_to_xlsx.py ──────────────────────────────────────────────
XSD_FILES = [                                                        # REUSED
    "SDSComXML.xsd", "SDSComXMLCT.xsd", "SDSComXMLDT.xsd",
    "SDSComXMLDT_GHS.xsd",
    "SDSComXMLNE_AR.xsd", "SDSComXMLNE_AT.xsd", "SDSComXMLNE_BR.xsd",
    "SDSComXMLNE_CA.xsd", "SDSComXMLNE_CH.xsd", "SDSComXMLNE_CN.xsd",
    "SDSComXMLNE_DE.xsd", "SDSComXMLNE_DK.xsd", "SDSComXMLNE_EU.xsd",
    "SDSComXMLNE_GB.xsd", "SDSComXMLNE_JP.xsd", "SDSComXMLNE_MX.xsd",
    "SDSComXMLNE_NO.xsd", "SDSComXMLNE_RU.xsd", "SDSComXMLNE_TR.xsd",
    "SDSComXMLNE_US.xsd",
]
XS_NS = "http://www.w3.org/2001/XMLSchema"                           # REUSED
XS_BUILTINS = {                                                      # REUSED
    "string","boolean","decimal","float","double","duration","dateTime",
    "time","date","gYearMonth","gYear","gMonthDay","gDay","gMonth",
    "hexBinary","base64Binary","anyURI","QName","normalizedString","token",
    "language","Name","NCName","ID","IDREF","IDREFS","NMTOKEN","NMTOKENS",
    "integer","nonPositiveInteger","negativeInteger","long","int","short",
    "byte","nonNegativeInteger","unsignedLong","unsignedInt","unsignedShort",
    "unsignedByte","positiveInteger","anyType","anySimpleType",
}

# Maps xs: local name → (samm_char_class, xsd_datatype | None)
# None datatype = predefined samm-c: characteristic, reference directly
XS_TO_SAMM: dict = {
    "string":             ("samm-c:Text",         None),
    "normalizedString":   ("samm-c:Text",         None),
    "token":              ("samm-c:Text",         None),
    "language":           ("samm-c:Text",         None),
    "Name":               ("samm-c:Text",         None),
    "NCName":             ("samm-c:Text",         None),
    "anyURI":             ("samm-c:ResourcePath", None),
    "boolean":            ("samm-c:Boolean",      None),
    "dateTime":           ("samm-c:Timestamp",    None),
    "date":               ("samm:Characteristic", "xsd:date"),
    "time":               ("samm:Characteristic", "xsd:time"),
    "gYear":              ("samm:Characteristic", "xsd:gYear"),
    "gYearMonth":         ("samm:Characteristic", "xsd:gYearMonth"),
    "integer":            ("samm:Characteristic", "xsd:integer"),
    "int":                ("samm:Characteristic", "xsd:int"),
    "long":               ("samm:Characteristic", "xsd:long"),
    "decimal":            ("samm:Characteristic", "xsd:decimal"),
    "float":              ("samm:Characteristic", "xsd:float"),
    "double":             ("samm:Characteristic", "xsd:double"),
    "positiveInteger":    ("samm:Characteristic", "xsd:positiveInteger"),
    "nonNegativeInteger": ("samm:Characteristic", "xsd:nonNegativeInteger"),
    "base64Binary":       ("samm:Characteristic", "xsd:base64Binary"),
    "hexBinary":          ("samm:Characteristic", "xsd:hexBinary"),
    "duration":           ("samm:Characteristic", "xsd:duration"),
}
PREDEFINED_CHARS = frozenset(                  # samm-c: types needing no definition
    cls for cls, dt in XS_TO_SAMM.values() if dt is None
)
MAX_DEPTH = 40                                                       # REUSED


# ═══════════════════════════════════════════════════════════════════════════════
# ── REUSED FROM xsd_to_xlsx.py ──────────────────────────────────────────────
# Functions below are copied verbatim or with minor signature changes from
# xsd_to_xlsx.py.  Any modification is noted inline with # MODIFIED.
# ═══════════════════════════════════════════════════════════════════════════════

def xstag(n):    return f"{{{XS_NS}}}{n}"                           # REUSED
def localtag(t): return t.split("}")[-1] if "}" in t else t         # REUSED
def strippfx(n): return n.split(":",1)[1] if ":" in n else n        # REUSED

def get_annotation(node) -> tuple:                                   # REUSED
    ann = node.find(xstag("annotation"))
    if ann is None:
        return "", frozenset()
    doc = " ".join(
        (d.text or "").strip()
        for d in ann.findall(xstag("documentation"))
        if (d.text or "").strip()
    )
    appinfos = frozenset(
        (ai.text or "").strip()
        for ai in ann.findall(xstag("appinfo"))
        if (ai.text or "").strip()
    )
    return doc, appinfos

def ensure_xsd_files(xsd_dir: str) -> None:                         # REUSED
    os.makedirs(xsd_dir, exist_ok=True)
    for fname in XSD_FILES:
        dest = os.path.join(xsd_dir, fname)
        if not os.path.exists(dest):
            print(f"  Downloading {fname} …")
            try:
                urllib.request.urlretrieve(GITHUB_RAW + fname, dest)
            except Exception as exc:
                print(f"  WARNING – {exc}")

def parse_all_xsd(xsd_dir: str):                                     # REUSED
    """First-definition-wins across all XSD files."""
    ct, st, re_el, grp, ag, ga = {}, {}, {}, {}, {}, {}
    type_origin: dict = {}
    for fname in XSD_FILES:
        path = os.path.join(xsd_dir, fname)
        if not os.path.exists(path):
            continue
        try:
            tree = ET.parse(path)
        except ET.ParseError:
            continue
        for child in tree.getroot():
            tag  = localtag(child.tag)
            name = child.get("name")
            if not name:
                continue
            if   tag == "complexType"    and name not in ct:
                ct[name] = child;  type_origin[name] = ("complexType", fname)
            elif tag == "simpleType"     and name not in st:
                st[name] = child;  type_origin[name] = ("simpleType",  fname)
            elif tag == "element"        and name not in re_el:  re_el[name] = child
            elif tag == "group"          and name not in grp:    grp[name]   = child
            elif tag == "attributeGroup" and name not in ag:     ag[name]    = child
            elif tag == "attribute"      and name not in ga:     ga[name]    = child
    return ct, st, re_el, grp, ag, ga, type_origin

# ═══════════════════════════════════════════════════════════════════════════════
# NAME UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def normalize_iri(name: str) -> str:
    """
    Normalize a PascalCase name for SAMM/Catena-X IRI conventions
    (Google Java Style Guide §5.3 acronym handling: an acronym run is
    treated as a single word, e.g. XmlHttpRequest not XMLHTTPRequest).

      SDSComXML            → SdsComXml
      ClassificationEU     → ClassificationEu
      GHSData              → GhsData
      DataSheet            → DataSheet          (unchanged)
      TimeDifferenceGMTID  → TimeDifferenceGmtId  (trailing acronym-pair)
    """
    s = name
    s = re.sub(r'(?<=[a-z0-9])(?=[A-Z])', '_', s)   # lower/digit → upper boundary
    s = re.sub(r'(?<=[A-Z])ID$', '_ID', s)          # trailing "...ID" acronym special-case:
                                                     # two back-to-back all-caps acronyms at the
                                                     # very end (no lowercase anchor) are otherwise
                                                     # indistinguishable; "ID" is common enough in
                                                     # XSD schemas to warrant this explicit split
    s = re.sub(r'(?<=[A-Z])(?=[A-Z][a-z])', '_', s) # general acronym → word boundary
    return "".join(p.capitalize() for p in s.split('_') if p)

def prop_iri(elem_name: str, type_local: Optional[str] = None,
             multi_type_names: Optional[Set[str]] = None) -> str:
    """lowerCamelCase property IRI, with type suffix when deduplication needed."""
    norm = normalize_iri(elem_name)
    base = norm[0].lower() + norm[1:]
    if multi_type_names and elem_name in multi_type_names and type_local:
        return base + normalize_iri(type_local)
    return base

def entity_iri(name: str) -> str:
    return normalize_iri(name)

def pref_name(name: str) -> str:
    """PascalCase → spaced words for samm:preferredName."""
    norm = normalize_iri(name)
    words = re.sub(r'([a-z])([A-Z])', r'\1 \2', norm)
    return words.strip()

def esc(s: str) -> str:
    """Escape a string for a Turtle double-quoted literal."""
    return s.replace('\\','\\\\').replace('"','\\"').replace('\n','\\n').replace('\r','')

def default_description(kind: str, human_name: str, extra: str = "") -> str:
    """
    Produce a consistent, comprehensible fallback samm:description when the
    XSD provides no (or empty) xs:documentation for a model element.
    All fallback templates share one style: short noun phrase, present tense,
    ending with a period, referencing the eSDScom XML schema as the source.
    """
    templates = {
        "aspect":            "{n} data as defined by the eSDScom XML schema.",
        "property":          "{n} property from the eSDScom XML schema.",
        "entity":            "{n} structure from the eSDScom XML schema.",
        "entity_stub":       "{n} structure from the eSDScom XML schema "
                             "(kept minimal to avoid an unresolved reference cycle).",
        "single_entity_char":"Single {n} value.",
        "list_char":         "List of {n} values.",
        "enum_char":         "Allowed values for {n}.",
        "xsd_type_char":     "Value of XSD type {n}.",
    }
    template = templates.get(kind, "{n}.")
    text = template.format(n=human_name)
    if extra:
        text = text.rstrip(".") + f"; {extra}."
    return text

# Heuristic-only: flags text that is LIKELY not English so it can be reviewed
# and translated by a human. This is not a translator — automatic translation
# of schema documentation is out of scope and risks silently changing meaning.
_NON_ENGLISH_MARKERS = (
    "ä", "ö", "ü", "ß", "Ä", "Ö", "Ü",
    " und ", " oder ", " für ", " der ", " die ", " das ",
    " mit ", " bei ", " eine ", " einen ", " ist ", " nicht ",
    " wird ", " werden ", " kann ", " sowie ", " gemäß ",
)

def looks_non_english(text: str) -> bool:
    """Heuristic check: True if `text` likely contains German (or similar) words."""
    if not text:
        return False
    padded = f" {text} "
    return any(marker in padded for marker in _NON_ENGLISH_MARKERS)

# samm:exampleValue templates for scalar xsd: datatypes — used for every custom
# Characteristic whose samm:dataType is an xsd: type (Entity-referencing
# SingleEntity/List characteristics are excluded; their datatype is a local
# entity IRI, not xsd:, so no exampleValue applies there).
_EXAMPLE_VALUES: dict = {
    "xsd:string":             "Example text",
    "xsd:boolean":             "true",
    "xsd:date":                "2026-07-20",
    "xsd:time":                "13:45:00",
    "xsd:dateTime":            "2026-07-20T13:45:00",
    "xsd:gYear":               "2026",
    "xsd:gYearMonth":          "2026-07",
    "xsd:gMonthDay":           "--07-20",
    "xsd:gDay":                "---20",
    "xsd:gMonth":              "--07",
    "xsd:integer":             "42",
    "xsd:int":                 "42",
    "xsd:long":                "42",
    "xsd:short":               "42",
    "xsd:byte":                "1",
    "xsd:decimal":             "3.14",
    "xsd:float":               "3.14",
    "xsd:double":              "3.14",
    "xsd:positiveInteger":     "1",
    "xsd:nonNegativeInteger":  "0",
    "xsd:nonPositiveInteger":  "0",
    "xsd:negativeInteger":     "-1",
    "xsd:unsignedLong":        "42",
    "xsd:unsignedInt":         "42",
    "xsd:unsignedShort":       "42",
    "xsd:unsignedByte":        "1",
    "xsd:base64Binary":        "QUJD",
    "xsd:hexBinary":           "1A2B3C",
    "xsd:duration":            "P1DT0H0M0S",
    "xsd:anyURI":              "https://example.com",
}

def example_value_for(xsd_dt: str, enum_vals: Optional[list] = None) -> str:
    """
    Return a samm:exampleValue that semantically fits the given xsd: datatype.
    For Enumeration characteristics, the first allowed value is used (it is
    by definition a valid, semantically fitting example).
    """
    if enum_vals:
        return str(enum_vals[0])
    return _EXAMPLE_VALUES.get(xsd_dt, "Example value")

# Predefined samm-c: characteristics used by this pipeline (PREDEFINED_CHARS)
# are scalar too — each wraps an implicit xsd: datatype under the hood, just
# without a local samm:Characteristic definition. Map each to that datatype
# so a matching samm:exampleValue can still be attached to the Property.
_PREDEFINED_CHAR_DATATYPE: dict = {
    "samm-c:Text":         "xsd:string",
    "samm-c:Boolean":      "xsd:boolean",
    "samm-c:Timestamp":    "xsd:dateTime",
    "samm-c:ResourcePath": "xsd:anyURI",
}

def example_literal_for_property(char: str, model: dict) -> Optional[str]:
    """
    samm:exampleValue belongs on the Property, not the Characteristic (the
    SAMM meta-model has no exampleValue slot on Characteristic — putting it
    there fails Aspect Model Editor validation). This looks up, for a given
    samm:characteristic reference, the example value semantically fitting
    its underlying scalar datatype and returns a ready-to-use typed Turtle
    literal such as '"2026-07-20"^^xsd:date'.

    Handles both:
      - local custom characteristics (":FooCharacteristic", scalar or Enum)
      - predefined samm-c: characteristics (samm-c:Text, samm-c:Boolean,
        samm-c:Timestamp, samm-c:ResourcePath) — these are scalar too, just
        without a local definition, so their datatype is looked up directly

    Returns None for Entity-referencing characteristics (SingleEntity/List),
    for which no scalar example applies.
    """
    if char.startswith(":"):
        c = model["characteristics"].get(char[1:])
        if not c or c.get("example") is None:
            return None
        return f'"{esc(c["example"])}"^^{c["datatype"]}'

    # Predefined samm-c: characteristic
    xsd_dt = _PREDEFINED_CHAR_DATATYPE.get(char)
    if xsd_dt is None:
        return None   # not a scalar predefined characteristic we know of
    return f'"{esc(example_value_for(xsd_dt))}"^^{xsd_dt}'


# ═══════════════════════════════════════════════════════════════════════════════
# XSD TRAVERSAL ITERATORS  (new – not in xsd_to_xlsx.py)
# ═══════════════════════════════════════════════════════════════════════════════

def iter_elem_children(ct_node, ctx, _seen=None):
    """
    Yield (elem_node, is_optional, is_list) for every xs:element
    reachable within a complexType node (including base-type inheritance).
    """
    if _seen is None:
        _seen = frozenset()
    for child in ct_node:
        tag = localtag(child.tag)
        if tag in ("sequence", "all", "choice"):
            yield from _iter_particle(child, ctx, tag != "sequence")
        elif tag == "complexContent":
            yield from _iter_cc_elems(child, ctx, _seen)

def _iter_particle(particle, ctx, parent_opt=False):
    for child in particle:
        tag = localtag(child.tag)
        if tag == "element":
            mo  = int(child.get("minOccurs", 1))
            mx  = child.get("maxOccurs", "1")
            lst = (mx == "unbounded") or (mx.isdigit() and int(mx) > 1)
            opt = parent_opt or (mo == 0)
            ref = child.get("ref")
            if ref:
                act = ctx["root_elements"].get(strippfx(ref))
                if act is not None:
                    yield (act, opt, lst)
            else:
                yield (child, opt, lst)
        elif tag in ("sequence", "all"):
            yield from _iter_particle(child, ctx, parent_opt)
        elif tag == "choice":
            yield from _iter_particle(child, ctx, True)
        elif tag == "group":
            ref = child.get("ref")
            if ref:
                g = ctx["groups"].get(strippfx(ref))
                if g is not None:
                    for sub in g:
                        if localtag(sub.tag) in ("sequence","choice","all"):
                            yield from _iter_particle(sub, ctx, parent_opt)

def _iter_cc_elems(cc_node, ctx, _seen):
    for child in cc_node:
        if localtag(child.tag) not in ("extension","restriction"):
            continue
        base = child.get("base", "")
        if base:
            bl = strippfx(base)
            if bl in ctx["complex_types"] and bl not in _seen:
                yield from iter_elem_children(ctx["complex_types"][bl], ctx, _seen | {bl})
        for sub in child:
            if localtag(sub.tag) in ("sequence","choice","all"):
                yield from _iter_particle(sub, ctx, False)

def iter_attr_children(ct_node, ctx, _seen=None):
    """Yield xs:attribute nodes from a complexType (including inherited)."""
    if _seen is None:
        _seen = frozenset()
    for child in ct_node:
        tag = localtag(child.tag)
        if tag == "attribute":
            yield child
        elif tag == "attributeGroup":
            yield from _iter_attr_group(child.get("ref",""), ctx)
        elif tag == "complexContent":
            yield from _iter_cc_attrs(child, ctx, _seen)
        elif tag == "simpleContent":
            yield from _iter_sc_attrs(child, ctx)

def _iter_attr_group(ref, ctx):
    if not ref:
        return
    ag = ctx["attribute_groups"].get(strippfx(ref))
    if ag is None:
        return
    for child in ag:
        tag = localtag(child.tag)
        if tag == "attribute":
            yield child
        elif tag == "attributeGroup":
            yield from _iter_attr_group(child.get("ref",""), ctx)

def _iter_cc_attrs(cc_node, ctx, _seen):
    for child in cc_node:
        if localtag(child.tag) not in ("extension","restriction"):
            continue
        base = child.get("base","")
        if base:
            bl = strippfx(base)
            if bl in ctx["complex_types"] and bl not in _seen:
                yield from iter_attr_children(ctx["complex_types"][bl], ctx, _seen | {bl})
        for sub in child:
            if localtag(sub.tag) == "attribute":
                yield sub
            elif localtag(sub.tag) == "attributeGroup":
                yield from _iter_attr_group(sub.get("ref",""), ctx)

def _iter_sc_attrs(sc_node, ctx):
    for child in sc_node:
        if localtag(child.tag) in ("extension","restriction"):
            for sub in child:
                if localtag(sub.tag) == "attribute":
                    yield sub
                elif localtag(sub.tag) == "attributeGroup":
                    yield from _iter_attr_group(sub.get("ref",""), ctx)


# ═══════════════════════════════════════════════════════════════════════════════
# SAMM TYPE RESOLUTION
# ═══════════════════════════════════════════════════════════════════════════════

def resolve_simple_type(type_local: str, ctx: dict) -> tuple:
    """
    Return (samm_kind, xsd_datatype_or_None, enum_values_or_None).
    samm_kind        – e.g. "samm-c:Text", "samm:Characteristic", "samm-c:Enumeration"
    xsd_datatype     – e.g. "xsd:date" or None for predefined characteristics
    enum_values      – list of str for Enumeration, else None
    """
    if type_local in XS_TO_SAMM:
        cls, dt = XS_TO_SAMM[type_local]
        return cls, dt, None

    st = ctx["simple_types"].get(type_local)
    if st is None:
        return "samm-c:Text", None, None  # unknown → fallback

    restr = st.find(xstag("restriction"))
    union = st.find(xstag("union"))

    if restr is not None:
        base = strippfx(restr.get("base", ""))
        enum_vals = [e.get("value","") for e in restr.findall(xstag("enumeration"))
                     if e.get("value")]
        if enum_vals:
            # Enumeration: determine xsd datatype from base
            _, base_dt, _ = resolve_simple_type(base, ctx)
            xsd_dt = base_dt or f"xsd:{base}" if base in XS_BUILTINS else "xsd:string"
            # Ensure proper xsd: prefix
            if not str(xsd_dt).startswith("xsd:"):
                xsd_dt = "xsd:string"
            return "samm-c:Enumeration", xsd_dt, enum_vals
        # Plain restriction: delegate to base
        return resolve_simple_type(base, ctx)

    if union is not None:
        return "samm-c:Text", None, None  # union → string fallback

    return "samm-c:Text", None, None  # unknown → fallback


def resolve_char_ref(type_local: str, is_list: bool,
                     model: dict, ctx: dict, building: set, parent_xpath: str) -> str:
    """
    Return the complete TTL reference string for the SAMM characteristic of type_local.
    Side-effects: may add entries to model['characteristics'] and model['entities'].
    `parent_xpath` is the XPath of the element that references this type — it is
    threaded through to ensure_entity() so each Entity's own properties get the
    correct absolute XPath recorded on first (and only) construction.
      ":DataSheetList"      → model-local (List)
      ":DataSheetChar"      → model-local (SingleEntity)
      ":DateCharacteristic" → model-local (custom Characteristic)
      "samm-c:Text"         → predefined, no definition needed
    """
    if not type_local:
        return "samm-c:Text"

    norm = entity_iri(type_local)   # normalized PascalCase for IRI

    # ── Stop type ───────────────────────────────────────────────────────────
    if type_local in ctx["stop_types"]:
        _ensure_stub_entity(type_local, norm, model)
        ciri = norm + ("List" if is_list else "Characteristic")
        if ciri not in model["characteristics"]:
            kind_label = "list_char" if is_list else "single_entity_char"
            model["characteristics"][ciri] = {
                "kind":        "samm-c:List" if is_list else "samm-c:SingleEntity",
                "datatype":    f":{norm}",
                "description": default_description(kind_label, pref_name(type_local)),
            }
        return f":{ciri}"

    # ── Named complex type ───────────────────────────────────────────────────
    if type_local in ctx["complex_types"]:
        ensure_entity(type_local, model, ctx, building, parent_xpath)
        ciri = norm + ("List" if is_list else "Characteristic")
        if ciri not in model["characteristics"]:
            kind_label = "list_char" if is_list else "single_entity_char"
            model["characteristics"][ciri] = {
                "kind":        "samm-c:List" if is_list else "samm-c:SingleEntity",
                "datatype":    f":{norm}",
                "description": default_description(kind_label, pref_name(type_local)),
            }
        return f":{ciri}"

    # ── Simple / enumeration type ────────────────────────────────────────────
    kind, xsd_dt, enum_vals = resolve_simple_type(type_local, ctx)

    if xsd_dt is None and not enum_vals:
        return kind  # predefined (samm-c:Text, samm-c:Boolean, …)

    ciri = norm + "Characteristic"
    if ciri not in model["characteristics"]:
        desc_kind = "enum_char" if enum_vals else "xsd_type_char"
        desc_name = pref_name(type_local) if enum_vals else (xsd_dt or "xsd:string")
        entry: dict = {
            "kind":        kind,
            "datatype":    xsd_dt or "xsd:string",
            "description": default_description(desc_kind, desc_name),
            "example":     example_value_for(xsd_dt or "xsd:string", enum_vals),
        }
        if enum_vals:
            entry["values"] = enum_vals
        model["characteristics"][ciri] = entry
    return f":{ciri}"


# ═══════════════════════════════════════════════════════════════════════════════
# SAMM MODEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def collect_multi_type_names(ctx: dict) -> set:
    """
    Scan ALL complex types for element names used with more than one distinct type.
    Returns set of such element names → these need a type suffix for deduplication.
    """
    name_to_types: dict = {}
    for ct_node in ctx["complex_types"].values():
        for elem_node, _, _ in iter_elem_children(ct_node, ctx):
            name = elem_node.get("name")
            ts   = elem_node.get("type")
            if name and ts:
                name_to_types.setdefault(name, set()).add(strippfx(ts))
    return {n for n, types in name_to_types.items() if len(types) > 1}


def _ensure_stub_entity(type_local: str, norm_iri: str, model: dict) -> None:
    if norm_iri not in model["entities"]:
        model["entities"][norm_iri] = {
            "preferred_name": pref_name(type_local),
            "description":    default_description("entity_stub", pref_name(type_local)),
            "see":            [],
            "properties":     [],
        }


def _describe_complex_type(ct_node) -> str:
    """
    Produce a short human-readable summary of a complexType's immediate
    (and, for complexContent, one-level-nested) child structure — used to
    explain why a complexType yielded zero SAMM properties.
    Examples of output:
      "empty (no children)"
      "complexContent/extension base=Regulations (base has no own children; " \\
      "extension adds: xs:any)"
      "sequence contains only: xs:any"
      "attributes only, all use=prohibited"
    """
    children = list(ct_node)
    if not children:
        return "empty complexType (no children at all)"

    parts = []
    for child in children:
        tag = localtag(child.tag)
        if tag == "complexContent":
            cc_children = list(child)
            for cc in cc_children:
                cc_tag = localtag(cc.tag)
                if cc_tag in ("extension", "restriction"):
                    base = cc.get("base", "?")
                    inner = [localtag(c.tag) for c in cc]
                    inner_desc = ", ".join(f"xs:{t}" for t in inner) if inner else "nothing"
                    parts.append(
                        f"complexContent/{cc_tag} base={base} "
                        f"(own content: {inner_desc})"
                    )
        elif tag == "simpleContent":
            sc_children = list(child)
            for sc in sc_children:
                sc_tag = localtag(sc.tag)
                if sc_tag in ("extension", "restriction"):
                    base = sc.get("base", "?")
                    attrs = [c.get("name", c.get("ref", "?"))
                            for c in sc if localtag(c.tag) == "attribute"]
                    parts.append(
                        f"simpleContent/{sc_tag} base={base}"
                        + (f", attributes: {attrs}" if attrs else ", no attributes")
                    )
        elif tag in ("sequence", "choice", "all"):
            inner_tags = [localtag(c.tag) for c in child]
            if not inner_tags:
                parts.append(f"empty xs:{tag}")
            elif all(t == "any" for t in inner_tags):
                parts.append(f"xs:{tag} contains only xs:any wildcard(s)")
            else:
                parts.append(f"xs:{tag} with: {', '.join(inner_tags)}")
        elif tag == "attribute":
            use = child.get("use", "optional")
            parts.append(f"xs:attribute name={child.get('name', child.get('ref','?'))} use={use}")
        elif tag == "attributeGroup":
            parts.append(f"xs:attributeGroup ref={child.get('ref','?')}")
        elif tag == "annotation":
            continue  # not structurally relevant
        else:
            parts.append(f"xs:{tag}")

    return "; ".join(parts) if parts else "no structurally relevant children"


def ensure_entity(type_local: str, model: dict, ctx: dict, building: set,
                  parent_xpath: str) -> None:
    """
    Build a samm:Entity for type_local if not already present.
    Uses `building` set to detect and break cycles.
    `parent_xpath` is the absolute XPath of the element that first references
    this type; since an Entity is only ever built once (memoized), this is
    also the XPath under which every one of its properties gets recorded.
    """
    norm = entity_iri(type_local)

    # Defensive collision check: two DIFFERENT XSD type names normalizing to
    # the same IRI would otherwise be silently merged into one entity.
    norm_map = ctx.setdefault("norm_to_type", {})
    prior = norm_map.get(norm)
    if prior is not None and prior != type_local:
        ctx.setdefault("iri_collisions", []).append((norm, prior, type_local))
    else:
        norm_map[norm] = type_local

    if norm in model["entities"]:
        return
    if type_local in building:
        _ensure_stub_entity(type_local, norm, model)
        return

    building.add(type_local)
    ct_node = ctx["complex_types"].get(type_local)
    if ct_node is None:
        _ensure_stub_entity(type_local, norm, model)
        building.discard(type_local)
        return

    doc, appinfos = get_annotation(ct_node)
    prop_entries = []

    for elem_node, is_opt, is_list in iter_elem_children(ct_node, ctx):
        piri = _add_elem_property(elem_node, is_opt, is_list, model, ctx, building, parent_xpath)
        if piri:
            prop_entries.append((piri, is_opt))

    for attr_node in iter_attr_children(ct_node, ctx):
        piri = _add_attr_property(attr_node, model, ctx, parent_xpath)
        if piri:
            is_opt = (attr_node.get("use", "optional") != "required")
            prop_entries.append((piri, is_opt))

    # Diagnostic: if the complexType exists but yields zero properties,
    # record its raw child structure so main() can log *why* it's empty
    # (e.g. xs:any wildcard, empty extension, prohibited-only attributes).
    if not prop_entries:
        ctx.setdefault("empty_type_debug", {})[type_local] = \
            _describe_complex_type(ct_node)

    model["entities"][norm] = {
        "preferred_name": pref_name(type_local),
        "description":    doc.strip() if doc.strip() else default_description(
                              "entity", pref_name(type_local)),
        "see":            [f"{APPINFO_BASE}{ai}" for ai in sorted(appinfos)],
        "properties":     prop_entries,
    }
    building.discard(type_local)


def _add_elem_property(elem_node, is_opt: bool, is_list: bool,
                        model: dict, ctx: dict, building: set,
                        parent_xpath: str) -> Optional[str]:
    """
    Create (or reuse) a samm:Property for an xs:element node.
    Returns the property IRI, or None on failure.
    `parent_xpath` is the absolute XPath of the enclosing element; this
    element's own XPath (parent_xpath + "/" + name) is recorded in
    ctx["prop_xpath"] the first time each property IRI is created — later
    reuse of the same (deduplicated) property elsewhere in the schema does
    not overwrite it, matching "one row per exported property".
    """
    name = elem_node.get("name")
    if not name:
        return None

    xpath = f"{parent_xpath}/{name}"

    type_str   = elem_node.get("type")
    type_local = strippfx(type_str) if type_str else None
    inline_ct  = elem_node.find(xstag("complexType"))
    inline_st  = elem_node.find(xstag("simpleType"))
    doc, appinfos = get_annotation(elem_node)
    see_refs   = [f"{APPINFO_BASE}{ai}" for ai in sorted(appinfos)]

    piri = prop_iri(name, type_local, ctx.get("multi_type_names"))

    # ── Determine characteristic reference ──────────────────────────────────
    if type_str:
        char = resolve_char_ref(type_local, is_list, model, ctx, building, xpath)

    elif inline_ct is not None:
        # Anonymous complex type: entity named after the element
        anon_iri = entity_iri(name)
        if anon_iri not in model["entities"] and name not in building:
            building.add(name)
            inner_props = []
            for ie, io, il in iter_elem_children(inline_ct, ctx):
                ip = _add_elem_property(ie, io, il, model, ctx, building, xpath)
                if ip:
                    inner_props.append((ip, io))
            for ia in iter_attr_children(inline_ct, ctx):
                ip = _add_attr_property(ia, model, ctx, xpath)
                if ip:
                    inner_props.append((ip, ia.get("use","optional") != "required"))
            anon_doc, anon_ai = get_annotation(inline_ct)
            model["entities"][anon_iri] = {
                "preferred_name": pref_name(name),
                "description":    anon_doc.strip() if anon_doc.strip() else
                                  default_description("entity", pref_name(name)),
                "see":            [f"{APPINFO_BASE}{ai}" for ai in sorted(anon_ai)],
                "properties":     inner_props,
            }
            building.discard(name)
        ciri = anon_iri + ("List" if is_list else "Characteristic")
        if ciri not in model["characteristics"]:
            kind_label = "list_char" if is_list else "single_entity_char"
            model["characteristics"][ciri] = {
                "kind":        "samm-c:List" if is_list else "samm-c:SingleEntity",
                "datatype":    f":{anon_iri}",
                "description": default_description(kind_label, pref_name(name)),
            }
        char = f":{ciri}"

    elif inline_st is not None:
        kind, xsd_dt, enum_vals = _resolve_inline_st(inline_st, ctx)
        if xsd_dt is None and not enum_vals:
            char = kind
        else:
            ciri = entity_iri(name) + "InlineCharacteristic"
            if ciri not in model["characteristics"]:
                desc_kind = "enum_char" if enum_vals else "xsd_type_char"
                desc_name = pref_name(name) if enum_vals else (xsd_dt or "xsd:string")
                entry: dict = {
                    "kind":        kind,
                    "datatype":    xsd_dt or "xsd:string",
                    "description": default_description(desc_kind, desc_name),
                    "example":     example_value_for(xsd_dt or "xsd:string", enum_vals),
                }
                if enum_vals:
                    entry["values"] = enum_vals
                model["characteristics"][ciri] = entry
            char = f":{ciri}"
    else:
        char = "samm-c:Text"

    if piri not in model["properties"]:
        ctx.setdefault("prop_xpath", {})[piri] = xpath
        model["properties"][piri] = {
            "preferred_name":  pref_name(name),
            "description":     doc.strip() if doc.strip() else
                               default_description("property", pref_name(name)),
            "raw_description": doc.strip(),   # empty if the XSD had no xs:documentation
                                               # (used for the review XLSX, kept separate
                                               # from the TTL's always-filled description)
            "see":             see_refs,
            "characteristic":  char,
            "example_literal": example_literal_for_property(char, model),
        }
    return piri


def _add_attr_property(attr_node, model: dict, ctx: dict,
                       parent_xpath: str) -> Optional[str]:
    """
    Create (or reuse) a samm:Property for an xs:attribute node.
    Returns the property IRI, or None on failure.
    `parent_xpath` is the absolute XPath of the enclosing element; per XPath
    convention the attribute itself is addressed as parent_xpath/@name.
    """
    ref = attr_node.get("ref")
    if ref:
        actual = ctx["global_attributes"].get(strippfx(ref))
        return _add_attr_property(actual, model, ctx, parent_xpath) if actual is not None else None

    name = attr_node.get("name")
    if not name or attr_node.get("use","optional") == "prohibited":
        return None

    xpath = f"{parent_xpath}/@{name}"

    type_str   = attr_node.get("type")
    type_local = strippfx(type_str) if type_str else None
    inline_st  = attr_node.find(xstag("simpleType"))
    doc, appinfos = get_annotation(attr_node)
    see_refs   = [f"{APPINFO_BASE}{ai}" for ai in sorted(appinfos)]

    piri = prop_iri(name)   # attributes: no dedup suffix (always simple types)

    if type_str:
        kind, xsd_dt, enum_vals = resolve_simple_type(type_local, ctx)
        if xsd_dt is None and not enum_vals:
            char = kind
        else:
            ciri = entity_iri(type_local or name) + "Characteristic"
            if ciri not in model["characteristics"]:
                desc_kind = "enum_char" if enum_vals else "xsd_type_char"
                desc_name = pref_name(type_local or name) if enum_vals else (xsd_dt or "xsd:string")
                entry: dict = {
                    "kind":        kind,
                    "datatype":    xsd_dt or "xsd:string",
                    "description": default_description(desc_kind, desc_name),
                    "example":     example_value_for(xsd_dt or "xsd:string", enum_vals),
                }
                if enum_vals:
                    entry["values"] = enum_vals
                model["characteristics"][ciri] = entry
            char = f":{ciri}"
    elif inline_st is not None:
        kind, xsd_dt, enum_vals = _resolve_inline_st(inline_st, ctx)
        char = kind if (xsd_dt is None and not enum_vals) else "samm-c:Text"
    else:
        char = "samm-c:Text"

    if piri not in model["properties"]:
        ctx.setdefault("prop_xpath", {})[piri] = xpath
        model["properties"][piri] = {
            "preferred_name":  pref_name(name),
            "description":     doc.strip() if doc.strip() else
                               default_description("property", pref_name(name)),
            "raw_description": doc.strip(),
            "see":             see_refs,
            "characteristic":  char,
            "example_literal": example_literal_for_property(char, model),
        }
    return piri


def _resolve_inline_st(st_node, ctx: dict) -> tuple:
    restr = st_node.find(xstag("restriction"))
    union = st_node.find(xstag("union"))
    if restr is not None:
        base = strippfx(restr.get("base",""))
        return resolve_simple_type(base, ctx)
    if union is not None:
        return "samm-c:Text", None, None
    return "samm-c:Text", None, None


def build_model(ctx: dict) -> dict:
    """Entry point: build the complete SAMM model from parsed XSD context."""
    model: dict = {
        "aspect":          {},
        "properties":      {},
        "characteristics": {},
        "entities":        {},
    }
    building: set = set()

    root_elem = ctx["root_elements"].get("DatasheetFeed")
    if root_elem is None:
        raise ValueError("DatasheetFeed not found in any XSD file")

    doc, appinfos = get_annotation(root_elem)

    # Determine the content node for DatasheetFeed
    type_str  = root_elem.get("type")
    inline_ct = root_elem.find(xstag("complexType"))
    if type_str:
        ct_node = ctx["complex_types"].get(strippfx(type_str))
    elif inline_ct is not None:
        ct_node = inline_ct
    else:
        ct_node = None

    aspect_props: list = []
    building.add("DatasheetFeed")
    root_xpath = "//DatasheetFeed"   # the actual XML root tag name, independent
                                     # of the SAMM Aspect's own IRI (ASPECT_IRI)
    if ct_node is not None:
        for elem_node, is_opt, is_list in iter_elem_children(ct_node, ctx):
            piri = _add_elem_property(elem_node, is_opt, is_list, model, ctx, building, root_xpath)
            if piri:
                aspect_props.append((piri, is_opt))
        for attr_node in iter_attr_children(ct_node, ctx):
            piri = _add_attr_property(attr_node, model, ctx, root_xpath)
            if piri:
                aspect_props.append((piri, attr_node.get("use","optional") != "required"))
    building.discard("DatasheetFeed")

    model["aspect"] = {
        "iri":            ASPECT_IRI,
        "preferred_name": "Datasheet Feed",
        "description":    doc.strip() if doc.strip() else
                          default_description("aspect", "Datasheet Feed"),
        "see":            [ASPECT_SEE_URL] + [f"{APPINFO_BASE}{ai}" for ai in sorted(appinfos)
                                              if f"{APPINFO_BASE}{ai}" != ASPECT_SEE_URL],
        "properties":     aspect_props,
    }
    return model


# ═══════════════════════════════════════════════════════════════════════════════
# TTL SERIALIZER
# ═══════════════════════════════════════════════════════════════════════════════

def _props_list(entries: list, props: dict) -> str:
    """
    Format samm:properties RDF list.
    entries: [(prop_iri, is_optional), …]
    Optional properties: [ samm:property :p ; samm:optional true ]
    """
    if not entries:
        return "( )"
    parts = []
    for piri, is_opt in entries:
        if is_opt:
            parts.append(f"[ samm:property :{piri} ; samm:optional true ]")
        else:
            parts.append(f":{piri}")
    if len(parts) == 1:
        return f"( {parts[0]} )"
    indent = "\n        "
    return "(" + indent + (indent).join(parts) + "\n    )"


def _see_line(see_refs: list) -> str:
    if not see_refs:
        return ""
    return "    samm:see " + ", ".join(f"<{r}>" for r in see_refs) + " ;"


def filter_model(model: dict) -> tuple:
    """
    Analyse the model and return a filtered copy plus lists for logging.

    Removing a "dangling" element can make something *else* dangling in turn:
      empty entity → characteristic pointing to it becomes orphaned
      orphaned characteristic → property using it becomes dangling
      property removed from an entity's samm:properties → that entity may
        itself become empty → loop

    This function runs a fixed-point loop so ALL of these are resolved
    together; nothing dangling is left in the returned model.

    Returns
    -------
    filtered_model      – copy of model with every unused/dangling element
                          removed (entities, characteristics, properties)
    skipped_chars       – custom characteristics never referenced by any
                          property (independent of entity emptiness)
    skipped_entities    – entities whose samm:properties list is empty
                          (either from the start, or after cascading)
    orphaned_chars      – characteristics that WERE referenced by a property,
                          but were removed because their samm:dataType points
                          to a (now) skipped entity
    skipped_properties  – properties removed because their samm:characteristic
                          was an orphaned_chars entry (cascading consequence)
    """
    # Work on copies so the caller's raw model is left untouched.
    entities = {k: {**v, "properties": list(v["properties"])}
                for k, v in model["entities"].items()}
    properties = dict(model["properties"])
    characteristics = dict(model["characteristics"])
    aspect = {**model["aspect"], "properties": list(model["aspect"]["properties"])}

    # ── Baseline: characteristics never referenced by any property at all ─────
    # (computed once, on the raw model — independent of entity-emptiness cascade)
    raw_used_by_prop = {
        p["characteristic"][1:]
        for p in properties.values()
        if p["characteristic"].startswith(":")
    }
    skipped_chars_unused = {k for k in characteristics if k not in raw_used_by_prop}

    skipped_entities:   set = set()
    skipped_chars:      set = set(skipped_chars_unused)
    skipped_properties: set = set()

    changed = True
    while changed:
        changed = False

        # 1. Entities whose (current) properties list is empty
        for eiri, e in entities.items():
            if eiri not in skipped_entities and not e["properties"]:
                skipped_entities.add(eiri)
                changed = True

        # 2. Characteristics pointing at a now-skipped entity → orphaned
        absent_dt = {f":{e}" for e in skipped_entities}
        for ciri, c in characteristics.items():
            if ciri not in skipped_chars and c.get("datatype", "") in absent_dt:
                skipped_chars.add(ciri)
                changed = True

        # 3. Properties whose characteristic is now skipped → dangling, remove
        for piri, p in properties.items():
            if piri in skipped_properties:
                continue
            char = p["characteristic"]
            if char.startswith(":") and char[1:] in skipped_chars:
                skipped_properties.add(piri)
                changed = True

        # 4. Remove skipped properties from every entity's and the aspect's
        #    samm:properties list (this may empty out further entities → loop)
        for e in entities.values():
            new_props = [(pi, o) for pi, o in e["properties"] if pi not in skipped_properties]
            if len(new_props) != len(e["properties"]):
                e["properties"] = new_props
                changed = True
        new_aspect_props = [(pi, o) for pi, o in aspect["properties"] if pi not in skipped_properties]
        if len(new_aspect_props) != len(aspect["properties"]):
            aspect["properties"] = new_aspect_props
            changed = True

    orphaned_chars = sorted(skipped_chars - skipped_chars_unused)

    filtered: dict = {
        "aspect":          aspect,
        "properties":      {k: v for k, v in properties.items()
                            if k not in skipped_properties},
        "characteristics": {k: v for k, v in characteristics.items()
                            if k not in skipped_chars},
        "entities":        {k: v for k, v in entities.items()
                            if k not in skipped_entities},
    }

    return (filtered,
            sorted(skipped_chars_unused),
            sorted(skipped_entities),
            orphaned_chars,
            sorted(skipped_properties))


def validate_model(model: dict) -> dict:
    """
    Audit the (filtered) model against the SAMM/Catena-X naming and
    documentation conventions requested for review:

      1. CamelCase (Google Java Style Guide §5.3 acronym handling)
      2. No "__" (double underscore) anywhere in an identifier
      3. Model elements (Aspect/Entity/Characteristic) start with a capital letter
      4. Properties start with a lowercase letter
      5. Every model element has a non-empty preferredName AND description,
         in English (heuristic check — flagged for manual review, not auto-fixed)
      6. A Property and its referenced Characteristic must not share the same name
      7. Every Property has a samm:exampleValue — checked only for properties
         whose characteristic is scalar (predefined samm-c:* or a local
         samm:Characteristic/samm-c:Enumeration). Entity-referencing properties
         (SingleEntity/List) are exempt: a literal exampleValue is not
         semantically meaningful for a whole nested structure.

    Returns a dict of issue-lists, one key per criterion; empty list = pass.
    """
    issues: dict = {
        "camel_case":         [],
        "double_underscore":  [],
        "capitalization":     [],
        "missing_fields":     [],
        "non_english":        [],
        "prop_char_same_name": [],
        "missing_example":    [],
    }

    CAMEL_SHAPE = re.compile(r'^[A-Za-z][A-Za-z0-9]*$')

    def check_identifier(iri: str, expect_upper: bool, where: str) -> None:
        if "__" in iri:
            issues["double_underscore"].append(f"{where} :{iri}")
        if not CAMEL_SHAPE.match(iri):
            issues["camel_case"].append(f"{where} :{iri}")
            return   # capitalization is meaningless if the shape itself is broken
        if expect_upper and not iri[0].isupper():
            issues["capitalization"].append(f"{where} :{iri} (must start with a capital letter)")
        elif not expect_upper and not iri[0].islower():
            issues["capitalization"].append(f"{where} :{iri} (must start with a lowercase letter)")

    def check_fields(preferred_name: str, description: str, where: str) -> None:
        if not preferred_name or not preferred_name.strip():
            issues["missing_fields"].append(f"{where}: missing samm:preferredName")
        if not description or not description.strip():
            issues["missing_fields"].append(f"{where}: missing samm:description")
        for label, text in (("preferredName", preferred_name), ("description", description)):
            if text and looks_non_english(text):
                issues["non_english"].append(
                    f"{where}.{label}: {text[:80]!r} — looks non-English, please review/translate"
                )

    def is_entity_referencing(char: str) -> bool:
        """True for SingleEntity/List characteristics (datatype = local Entity IRI),
        for which a scalar exampleValue is not applicable and thus not required."""
        if not char.startswith(":"):
            return False   # predefined samm-c: characteristic — always scalar
        c = model["characteristics"].get(char[1:])
        return bool(c) and c.get("kind") in ("samm-c:SingleEntity", "samm-c:List")

    # ── Aspect ─────────────────────────────────────────────────────────────
    a = model["aspect"]
    check_identifier(a["iri"], True, "Aspect")
    check_fields(a.get("preferred_name", ""), a.get("description", ""), f"Aspect :{a['iri']}")

    # ── Properties ─────────────────────────────────────────────────────────
    for piri, p in model["properties"].items():
        check_identifier(piri, False, "Property")
        check_fields(p.get("preferred_name", ""), p.get("description", ""), f"Property :{piri}")
        char = p.get("characteristic", "")
        if char.startswith(":"):
            char_local = char[1:]
            if char_local.lower() == piri.lower():
                issues["prop_char_same_name"].append(
                    f"Property :{piri} and its Characteristic :{char_local} share the same name"
                )
        if not is_entity_referencing(char) and not p.get("example_literal"):
            issues["missing_example"].append(f"Property :{piri}: missing samm:exampleValue")

    # ── Characteristics (custom only; predefined samm-c: types are out of our control) ──
    for ciri, c in model["characteristics"].items():
        check_identifier(ciri, True, "Characteristic")
        check_fields(pref_name(ciri), c.get("description", ""), f"Characteristic :{ciri}")

    # ── Entities ───────────────────────────────────────────────────────────
    for eiri, e in model["entities"].items():
        check_identifier(eiri, True, "Entity")
        check_fields(e.get("preferred_name", ""), e.get("description", ""), f"Entity :{eiri}")

    return issues


def print_validation_report(issues: dict) -> bool:
    """Print a PASS/FAIL report for each criterion. Returns True iff everything passed."""
    labels = {
        "camel_case":          "1. CamelCase (Google style)",
        "double_underscore":   "2. No double underscore ('__')",
        "capitalization":      "3./4. Capitalization (elements upper / properties lower)",
        "missing_fields":      "5a. preferredName + description present",
        "non_english":         "5b. English language (heuristic — manual review)",
        "prop_char_same_name": "6. Property/Characteristic name distinctness",
        "missing_example":     "7. exampleValue present on scalar properties",
    }
    all_ok = True
    print("\n  Naming & documentation convention check:")
    print("  " + "─" * 58)
    for key, label in labels.items():
        problems = issues[key]
        if not problems:
            print(f"    ✓ {label}: PASS")
        else:
            # "non_english" is advisory (heuristic), everything else is a hard fail
            severity = "REVIEW" if key == "non_english" else "FAIL"
            if severity == "FAIL":
                all_ok = False
            print(f"    {'⚠' if severity=='REVIEW' else '✗'} {label}: {severity} "
                  f"({len(problems)} issue(s))")
            for p in problems[:15]:
                print(f"        - {p}")
            if len(problems) > 15:
                print(f"        … and {len(problems) - 15} more")
    return all_ok


def write_ttl(model: dict, output_file: str) -> None:
    SV = SAMM_VERSION
    lines: list = []

    # ── Copyright header ───────────────────────────────────────────────────────
    lines += [
        "#" * 71,
        "# Copyright(c) 2026 Qualisys GmbH",
        "# Copyright(c) 2026 BASF SE",
        "# Copyright(c) 2026 Contributors to the Eclipse Foundation",
        "#",
        "# See the NOTICE file(s) distributed with this work for additional",
        "# information regarding copyright ownership.",
        "#",
        "# This work is made available under the terms of the",
        "# Creative Commons Attribution 4.0 International(CC-BY-4.0) license,",
        "# which is available at",
        "# https://creativecommons.org/licenses/by/4.0/legalcode.",
        "#",
        "# SPDX-License-Identifier: CC-BY-4.0",
        "#" * 71,
        "#",
        "# This ontology was generated from an XML schema definition by a",
        "# script and thus lacks the perfection of a native TTL file.",
        "#" * 71,
    ]

    # ── Prefixes ─────────────────────────────────────────────────────────────
    lines += [
        f"@prefix :       <{MODEL_NS}> .",
        f"@prefix samm:   <urn:samm:org.eclipse.esmf.samm:meta-model:{SV}#> .",
        f"@prefix samm-c: <urn:samm:org.eclipse.esmf.samm:characteristic:{SV}#> .",
        f"@prefix samm-e: <urn:samm:org.eclipse.esmf.samm:entity:{SV}#> .",
        f"@prefix unit:   <urn:samm:org.eclipse.esmf.samm:unit:{SV}#> .",
        "@prefix rdf:    <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .",
        "@prefix rdfs:   <http://www.w3.org/2000/01/rdf-schema#> .",
        "@prefix xsd:    <http://www.w3.org/2001/XMLSchema#> .",
        "",
    ]

    # ── Aspect ───────────────────────────────────────────────────────────────
    a     = model["aspect"]
    plist = _props_list(a["properties"], model["properties"])
    lines += [
        "# " + "─" * 74,
        "# Aspect",
        "# " + "─" * 74,
        "",
        f":{a['iri']} a samm:Aspect ;",
        f'    samm:preferredName "{esc(a["preferred_name"])}"@en ;',
        f'    samm:description "{esc(a["description"])}"@en ;',
    ]
    if a.get("see"):
        lines.append(_see_line(a["see"]))
    lines += [
        f"    samm:properties {plist} ;",
        "    samm:operations ( ) ;",
        "    samm:events ( ) .",
        "",
    ]

    # ── Properties ───────────────────────────────────────────────────────────
    lines += ["# " + "─" * 74, "# Properties", "# " + "─" * 74, ""]
    for piri in sorted(model["properties"]):
        p = model["properties"][piri]
        char = p["characteristic"]
        lines.append(f":{piri} a samm:Property ;")
        lines.append(f'    samm:preferredName "{esc(p["preferred_name"])}"@en ;')
        lines.append(f'    samm:description "{esc(p["description"])}"@en ;')
        sl = _see_line(p.get("see", []))
        if sl:
            lines.append(sl)
        if p.get("example_literal"):
            lines.append(f"    samm:exampleValue {p['example_literal']} ;")
        lines.append(f"    samm:characteristic {char} .")
        lines.append("")

    # ── Characteristics (custom only) ─────────────────────────────────────────
    custom_chars = {k: v for k, v in model["characteristics"].items()
                    if v["kind"] not in PREDEFINED_CHARS}
    if custom_chars:
        lines += ["# " + "─" * 74, "# Characteristics", "# " + "─" * 74, ""]
        for ciri in sorted(custom_chars):
            c = custom_chars[ciri]
            desc = c.get("description") or default_description("xsd_type_char", ciri)
            # Note: samm:exampleValue is NOT written here — the SAMM meta-model has
            # no exampleValue slot on Characteristic. It is instead attached to
            # every Property that uses this characteristic (see Properties section).
            if c["kind"] == "samm-c:Enumeration":
                vals = " ".join(f'"{esc(str(v))}"' for v in c.get("values", []))
                lines.append(f":{ciri} a samm-c:Enumeration ;")
                lines.append(f'    samm:preferredName "{esc(pref_name(ciri))}"@en ;')
                lines.append(f'    samm:description "{esc(desc)}"@en ;')
                lines.append(f"    samm:dataType {c['datatype']} ;")
                lines.append(f"    samm-c:values ( {vals} ) .")
            else:
                lines.append(f":{ciri} a {c['kind']} ;")
                lines.append(f'    samm:preferredName "{esc(pref_name(ciri))}"@en ;')
                lines.append(f'    samm:description "{esc(desc)}"@en ;')
                lines.append(f"    samm:dataType {c['datatype']} .")
            lines.append("")

    # ── Entities ─────────────────────────────────────────────────────────────
    lines += ["# " + "─" * 74, "# Entities", "# " + "─" * 74, ""]
    for eiri in sorted(model["entities"]):
        e = model["entities"][eiri]
        plist = _props_list(e["properties"], model["properties"])
        lines.append(f":{eiri} a samm:Entity ;")
        lines.append(f'    samm:preferredName "{esc(e["preferred_name"])}"@en ;')
        lines.append(f'    samm:description "{esc(e["description"])}"@en ;')
        sl = _see_line(e.get("see", []))
        if sl:
            lines.append(sl)
        lines.append(f"    samm:properties {plist} .")
        lines.append("")

    content = "\n".join(lines)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(content)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def write_property_xlsx(model: dict, prop_xpath: dict, output_file: str) -> None:
    """
    Write an XLSX with one row per exported SAMM Property:
      A  absolute XPath of the XSD element/attribute that produced it
         (e.g. //DatasheetFeed/DataSheet/InformationFromExportingSystem)
      B  SAMM Property local name
      C  SAMM Description (may be empty)

    Every property in model["properties"] is listed — including those with
    no XSD documentation at all, for which column C is left empty so a
    reviewer can immediately spot which properties still need documentation.
    (The TTL itself always carries a non-empty samm:description via a
    synthesized fallback; that fallback text is intentionally NOT shown
    here, since it would hide the fact that the source XSD had none.)

    `prop_xpath` maps property-IRI → XPath, recorded at first construction
    during build_model(). A property may in principle originate from a type
    reused at several XPath locations (e.g. a shared Entity) — the FIRST
    location encountered during traversal is used, keeping exactly one row
    per property as requested.
    """
    rows = []
    for piri, p in model["properties"].items():
        xpath = prop_xpath.get(piri, "")
        raw_desc = p.get("raw_description", p.get("description", ""))
        rows.append((xpath, piri, raw_desc))
    rows.sort(key=lambda r: r[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties"

    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F3864")
    hdr_align = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    for col, header in enumerate(("XPath", "SAMM Property", "SAMM Description"), 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, hdr_align

    mono_font = Font(name="Courier New", size=9)
    data_font = Font(name="Arial", size=10)
    left_align = Alignment(horizontal="left", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i, (xpath, piri, desc) in enumerate(rows, 2):
        c1 = ws.cell(row=i, column=1, value=xpath)
        c2 = ws.cell(row=i, column=2, value=piri)
        c3 = ws.cell(row=i, column=3, value=desc)
        c1.font, c1.alignment = mono_font, left_align
        c2.font, c2.alignment = data_font, left_align
        c3.font, c3.alignment = data_font, wrap_align

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(rows) + 1}"

    wb.save(output_file)


def main() -> None:
    xsd_dir     = sys.argv[1] if len(sys.argv) > 1 else "xsd_files"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "DatasheetFeed.ttl"
    xlsx_file   = sys.argv[3] if len(sys.argv) > 3 else \
                  os.path.splitext(output_file)[0] + "_properties.xlsx"

    print("=" * 62)
    print("eSDScom XSD → SAMM 2.1.0 TTL")
    print("=" * 62)
    print(f"XSD directory : {os.path.abspath(xsd_dir)}")
    print(f"Output file   : {os.path.abspath(output_file)}")
    print(f"Properties XLSX: {os.path.abspath(xlsx_file)}")
    print(f"Model NS      : {MODEL_NS}")
    print(f"Stop types    : {sorted(STOP_TYPES)}")

    print("\n[1/6] Ensuring XSD files …")
    ensure_xsd_files(xsd_dir)

    print("\n[2/6] Parsing XSD files …")
    ct, st, re_el, grp, ag, ga, type_origin = parse_all_xsd(xsd_dir)
    print(f"  Complex types     : {len(ct)}")
    print(f"  Simple types      : {len(st)}")
    print(f"  Root elements     : {len(re_el)}")
    print(f"  Attribute groups  : {len(ag)}")

    if "DatasheetFeed" not in re_el:
        print("ERROR: DatasheetFeed not found.")
        sys.exit(1)

    ctx = {
        "complex_types":     ct,
        "simple_types":      st,
        "root_elements":     re_el,
        "groups":            grp,
        "attribute_groups":  ag,
        "global_attributes": ga,
        "stop_types":        STOP_TYPES,
    }

    print("\n[3/6] Pre-pass: collecting multi-type element names …")
    mtn = collect_multi_type_names(ctx)
    ctx["multi_type_names"] = mtn
    print(f"  Names needing type suffix: {len(mtn)}")
    for n in sorted(mtn):
        print(f"    {n}")

    print("\n[4/6] Building SAMM model …")
    model = build_model(ctx)
    print(f"  Entities (raw)        : {len(model['entities'])}")
    print(f"  Properties (raw)      : {len(model['properties'])}")
    print(f"  Characteristics (raw) : {len(model['characteristics'])}")

    collisions = ctx.get("iri_collisions", [])
    if collisions:
        print(f"\n  WARNING – {len(collisions)} IRI collision(s): distinct XSD types "
              f"normalized to the same identifier (only the first was kept as an Entity):")
        print("  " + "─" * 58)
        for norm, first, second in collisions:
            print(f"    :{norm}  ←  {first!r} (kept) vs. {second!r} (dropped)")

    # ── Filter unused characteristics, empty entities, and cascading dangling
    #    properties (fixed-point: removing one can create another) ────────────
    raw_characteristics = model["characteristics"]   # kept for logging lookups below
    model, skipped_chars, skipped_entities, orphaned_chars, skipped_properties = \
        filter_model(model)

    SEP = "  " + "─" * 58
    debug = ctx.get("empty_type_debug", {})

    def _entity_reason(entity_name: str):
        reason = debug.get(entity_name)
        if reason is None:
            for orig, desc in debug.items():
                if entity_iri(orig) == entity_name:
                    return desc
        return reason

    if skipped_chars:
        print(f"\n  {len(skipped_chars)} unused characteristic(s) – skipped from TTL:")
        print(SEP)
        for c in skipped_chars:
            print(f"    :{c}")

    if skipped_entities:
        print(f"\n  {len(skipped_entities)} empty entity(ies) – skipped from TTL:")
        print(SEP)
        for e in skipped_entities:
            print(f"    :{e}")
            reason = _entity_reason(e)
            if reason:
                print(f"        └─ reason: {reason}")

    if orphaned_chars:
        print(f"\n  {len(orphaned_chars)} characteristic(s) referenced a skipped entity")
        print(f"  – skipped from TTL (would otherwise be dangling):")
        print(SEP)
        for c in orphaned_chars:
            dt = raw_characteristics.get(c, {}).get("datatype", "?")
            print(f"    :{c}  →  samm:dataType {dt}")
            entity_name = dt[1:] if dt.startswith(":") else dt
            reason = _entity_reason(entity_name)
            if reason:
                print(f"        └─ {dt} is empty because: {reason}")

    if skipped_properties:
        print(f"\n  {len(skipped_properties)} propert(y/ies) removed as a consequence")
        print(f"  (their samm:characteristic was one of the entries above):")
        print(SEP)
        for p in skipped_properties:
            print(f"    :{p}")

    print(f"\n  Entities written        : {len(model['entities'])}")
    print(f"  Properties written      : {len(model['properties'])}")
    print(f"  Characteristics written : {len(model['characteristics'])}")

    # ── Self-check: dangling references in the filtered model ─────────────────
    dangling = []
    for ap, _ in model["aspect"]["properties"]:
        if ap not in model["properties"]:
            dangling.append(f"Aspect :{model['aspect']['iri']} → :{ap} (property absent)")
    for piri, p in model["properties"].items():
        char = p["characteristic"]
        if char.startswith(":") and char[1:] not in model["characteristics"]:
            dangling.append(f"Property :{piri} → {char} (characteristic absent)")
    for eiri, e in model["entities"].items():
        for ep, _ in e["properties"]:
            if ep not in model["properties"]:
                dangling.append(f"Entity :{eiri} → :{ep} (property absent)")
    for ciri, c in model["characteristics"].items():
        dt = c.get("datatype", "")
        if dt.startswith(":") and dt[1:] not in model["entities"]:
            dangling.append(f"Characteristic :{ciri} → {dt} (entity absent)")

    if dangling:
        print(f"\n  SELF-CHECK – {len(dangling)} dangling reference(s) in output model:")
        print(SEP)
        for d in dangling[:30]:
            print(f"    {d}")
        if len(dangling) > 30:
            print(f"    … and {len(dangling)-30} more")
    else:
        print("  Self-check: all references resolve ✓")

    validation_issues = validate_model(model)
    print_validation_report(validation_issues)

    print("\n[5/6] Writing TTL …")
    write_ttl(model, output_file)
    print(f"  Saved: {output_file}")
    print(f"  Lines: {sum(1 for _ in open(output_file))}")

    print("\n[6/6] Writing property XLSX …")
    write_property_xlsx(model, ctx.get("prop_xpath", {}), xlsx_file)
    print(f"  Saved: {xlsx_file}")
    print(f"  Rows: {len(model['properties'])}")

    print("\nDone.")


if __name__ == "__main__":
    main()
